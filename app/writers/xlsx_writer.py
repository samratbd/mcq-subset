"""XLSX writer (same schema as the CSV writer)."""

from __future__ import annotations
import io
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from ..models import Question
from ..math_utils import render_text


HEADER = ["title", "type", "option1", "option2", "option3", "option4",
          "answer", "explanation"]


def write_xlsx(questions: List[Question], *, math_mode: str = "katex") -> bytes:
    if math_mode not in ("katex", "unicode"):
        raise ValueError(f"unknown math_mode for XLSX: {math_mode!r}")

    def m(s: str) -> str:
        return render_text(s, math_mode)

    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"

    ws.append(HEADER)
    for col_idx in range(1, len(HEADER) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top")

    for q in questions:
        ws.append([
            m(q.question),
            "MCQ",
            m(q.options[0]), m(q.options[1]), m(q.options[2]), m(q.options[3]),
            q.answer_index + 1,
            m(q.explanation),
        ])

    widths = [60, 8, 28, 28, 28, 28, 8, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
