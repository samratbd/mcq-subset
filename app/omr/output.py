"""Output writers for OMR scan results.

One row per scanned sheet. Columns:
    serial, roll_number, set, Q1, Q2, ..., QN, confidence, needs_review, review_items
"""

from __future__ import annotations
import csv
import io
import json
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .scanner import OmrResult


# Header row builder — used by all three formats.

def _headers(n_questions: int) -> List[str]:
    return (
        ["serial", "roll_number", "set"]
        + [f"Q{i + 1}" for i in range(n_questions)]
        + ["confidence", "needs_review", "review_items", "source_file"]
    )


def _row_for_result(serial: int,
                    result: OmrResult,
                    n_questions: int,
                    source_file: str) -> list:
    """Pad/truncate answer list to n_questions and assemble a row."""
    answers = list(result.answers)
    if len(answers) < n_questions:
        answers += [""] * (n_questions - len(answers))
    elif len(answers) > n_questions:
        answers = answers[:n_questions]

    return [
        serial,
        result.roll_number,
        result.set_letter,
        *answers,
        round(result.confidence * 100, 1),  # display as a percentage
        "yes" if result.needs_review else "",
        ",".join(result.review_items),
        source_file,
    ]


# --- CSV ---------------------------------------------------------------------

def write_csv(results: List[tuple[OmrResult, str]],
              n_questions: int) -> bytes:
    """`results` is a list of (OmrResult, source_filename) tuples."""
    buf = io.StringIO()
    w = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    w.writerow(_headers(n_questions))
    for serial, (res, src) in enumerate(results, start=1):
        w.writerow(_row_for_result(serial, res, n_questions, src))
    return "\ufeff".encode("utf-8") + buf.getvalue().encode("utf-8")


# --- XLSX --------------------------------------------------------------------

REVIEW_FILL = PatternFill(start_color="FFE6CC", end_color="FFE6CC",
                          fill_type="solid")        # light orange
EMPTY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2",
                         fill_type="solid")          # light grey


def write_xlsx(results: List[tuple[OmrResult, str]],
               n_questions: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "OMR Results"

    headers = _headers(n_questions)
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze the header row + the first three columns (serial / roll / set)
    ws.freeze_panes = "D2"

    for serial, (res, src) in enumerate(results, start=1):
        row_data = _row_for_result(serial, res, n_questions, src)
        ws.append(row_data)

        excel_row = serial + 1
        # Highlight the whole row if needs_review
        if res.needs_review:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=excel_row, column=col_idx).fill = REVIEW_FILL

        # Mark blank-answer cells in light grey so they stand out
        flagged_qs = {
            int(item[1:]) for item in res.review_items if item.startswith("Q")
        }
        for q_idx in range(n_questions):
            ans_col = 4 + q_idx  # serial(1) + roll(2) + set(3) + Qi(4..)
            cell = ws.cell(row=excel_row, column=ans_col)
            if not cell.value:
                cell.fill = EMPTY_FILL
            if (q_idx + 1) in flagged_qs:
                cell.font = Font(bold=True, color="CC3300")

    # Column widths
    widths = [6, 12, 6] + [6] * n_questions + [11, 12, 18, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Center-align answer cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=4, max_col=3 + n_questions):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- JSON --------------------------------------------------------------------

def write_json(results: List[tuple[OmrResult, str]],
               n_questions: int) -> bytes:
    out = {
        "n_questions": n_questions,
        "n_sheets": len(results),
        "sheets": [],
    }
    for serial, (res, src) in enumerate(results, start=1):
        d = res.as_dict()
        d["serial"] = serial
        d["source_file"] = src
        out["sheets"].append(d)
    return json.dumps(out, ensure_ascii=False, indent=2).encode("utf-8")
