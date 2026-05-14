"""XLSX parser.

Same schema as CSV (single sheet, header on row 1):
    title, type, option1, option2, option3, option4, answer, explanation
"""

from __future__ import annotations
import io
from typing import List

from openpyxl import load_workbook

from ..models import Question, num_to_idx


_HEADERS = ["title", "type", "option1", "option2", "option3", "option4",
            "answer", "explanation"]


def _norm(s) -> str:
    if s is None:
        return ""
    return str(s).strip().lstrip("\ufeff").lower()


def parse_xlsx_bytes(data: bytes) -> List[Question]:
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("XLSX has no rows")

    header = [_norm(c) for c in rows[0]]
    col_idx = {}
    for h in _HEADERS:
        if h not in header:
            raise ValueError(
                f"XLSX missing required column '{h}'. Found: {header}"
            )
        col_idx[h] = header.index(h)

    questions: List[Question] = []
    for r_offset, row in enumerate(rows[1:], start=1):
        row = list(row)
        if all((c is None or str(c).strip() == "") for c in row):
            continue
        if len(row) < len(header):
            row = list(row) + [None] * (len(header) - len(row))

        def cell(name):
            v = row[col_idx[name]]
            return "" if v is None else str(v).strip()

        title = cell("title")
        opts = [cell(f"option{i}") for i in (1, 2, 3, 4)]
        ans_raw = cell("answer")
        expl = cell("explanation")

        if not title or any(o == "" for o in opts):
            raise ValueError(
                f"XLSX row {r_offset + 1}: title or one of option1..4 is empty"
            )

        try:
            answer_idx = num_to_idx(ans_raw)
        except ValueError as e:
            raise ValueError(f"XLSX row {r_offset + 1}: {e}") from None

        questions.append(Question(
            sl=r_offset,
            question=title,
            options=opts,
            answer_index=answer_idx,
            explanation=expl,
        ))

    if not questions:
        raise ValueError("XLSX had a header but no data rows")
    return questions
