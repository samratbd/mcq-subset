#!/usr/bin/env python3
"""Desktop OMR Scanner — single-window Tkinter app.

A standalone GUI for scanning OMR sheets without the web server. Built so
the user can:

  * pick a folder of scanned sheets (BMP / PNG / JPEG / TIFF)
  * choose 50-mark or 100-mark template (or auto-detect)
  * process the whole batch — hundreds or thousands of sheets
  * see live progress (per-sheet status, % done, ETA)
  * stop the run at any time
  * save results to XLSX, CSV, or JSON
  * optionally save annotated review images alongside

No web server, no upload size limits, no time-outs. Runs on the user's
local machine and uses every available CPU core.

Usage
-----
    python desktop_omr.py

Dependencies (the same as the web app):
    pip install opencv-python-headless numpy Pillow openpyxl

Build instructions for a standalone .exe / .app
-----------------------------------------------
PyInstaller turns this into a single double-clickable executable:

    pip install pyinstaller
    pyinstaller --onefile --windowed desktop_omr.py

Output:
    dist/desktop_omr.exe   (Windows)
    dist/desktop_omr.app   (macOS)
    dist/desktop_omr       (Linux)
"""

from __future__ import annotations

import csv
import json
import os
import queue
import re
import sys
import threading
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import List, Optional, Tuple

# Tkinter is in Python's stdlib — no install needed.
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# Add the bundled `app/` package to the path so we can reuse the same OMR code
# the web server uses. When packaged with PyInstaller, this path is inside
# the unpacked bundle (sys._MEIPASS).
_BASE_DIR = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
sys.path.insert(0, str(_BASE_DIR))

try:
    from app.omr import scan_omr, render_review_image, TEMPLATES
    from app.omr.scanner import OmrResult
except ImportError as e:
    print(
        "ERROR: cannot import OMR module.\n"
        f"  Reason: {e}\n"
        "  Fix: run from the mcq_shuffler folder, "
        "and install dependencies with `pip install -r requirements.txt`.",
        file=sys.stderr,
    )
    sys.exit(1)


# -----------------------------------------------------------------------------
# Worker — runs in a background thread so the GUI stays responsive
# -----------------------------------------------------------------------------


class ScanJob:
    """One run of the scanner over a list of files."""

    SUPPORTED_EXTS = (
        ".bmp", ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".gif",
    )

    def __init__(
        self,
        files: List[Path],
        sheet_type: str,
        output_dir: Path,
        output_format: str,
        save_review_images: bool,
        progress_queue: "queue.Queue[dict]",
        cancel_event: threading.Event,
    ) -> None:
        self.files = files
        self.sheet_type = sheet_type
        self.output_dir = output_dir
        self.output_format = output_format
        self.save_review_images = save_review_images
        self.progress_queue = progress_queue
        self.cancel_event = cancel_event
        self.results: List[Tuple[OmrResult, str]] = []
        self.started_at = time.time()

    def _scan_one(self, file_path: Path) -> Tuple[OmrResult, str, Optional[bytes]]:
        """Scan one file. Returns (result, filename, review_png_or_None)."""
        if self.cancel_event.is_set():
            res = OmrResult(
                sheet_type=("omr_50" if self.sheet_type == "auto"
                            else self.sheet_type),
                roll_number="?", set_letter="?", answers=[],
                confidence=0.0, needs_review=True,
                review_items=["cancelled"], fill_fractions=[],
                error="Cancelled by user",
            )
            return res, file_path.name, None
        try:
            data = file_path.read_bytes()
            res = scan_omr(data, sheet_type=self.sheet_type)
        except Exception as e:
            res = OmrResult(
                sheet_type=("omr_50" if self.sheet_type == "auto"
                            else self.sheet_type),
                roll_number="?", set_letter="?", answers=[],
                confidence=0.0, needs_review=True,
                review_items=["scan_failed"], fill_fractions=[],
                error=f"{type(e).__name__}: {e}",
            )
            return res, file_path.name, None
        review_png: Optional[bytes] = None
        if self.save_review_images and not res.error:
            try:
                review_png = render_review_image(data, res)
            except Exception:
                review_png = None
        return res, file_path.name, review_png

    def run(self) -> None:
        """Process all files in parallel, then write the output."""
        try:
            self.output_dir.mkdir(parents=True, exist_ok=True)
            if self.save_review_images:
                (self.output_dir / "review").mkdir(exist_ok=True)

            self.progress_queue.put({
                "type": "start",
                "total": len(self.files),
            })

            # Tune worker count to CPU. OpenCV releases the GIL during the
            # heavy image work, so threads scale roughly linearly.
            max_workers = min(8, max(2, (os.cpu_count() or 2)))

            done = 0
            with ThreadPoolExecutor(max_workers=max_workers) as pool:
                futures = {
                    pool.submit(self._scan_one, fp): fp
                    for fp in self.files
                }
                for fut in as_completed(futures):
                    if self.cancel_event.is_set():
                        for f in futures:
                            f.cancel()
                        break
                    fp = futures[fut]
                    try:
                        res, fname, review_png = fut.result()
                    except Exception as e:
                        # Should be already caught inside _scan_one but be safe
                        res = OmrResult(
                            sheet_type="omr_50",
                            roll_number="?", set_letter="?", answers=[],
                            confidence=0.0, needs_review=True,
                            review_items=["worker_crash"],
                            fill_fractions=[],
                            error=f"Worker exception: {e}",
                        )
                        fname = fp.name
                        review_png = None

                    self.results.append((res, fname))

                    if review_png:
                        safe = re.sub(r"[^A-Za-z0-9._-]+", "_", fname)
                        (self.output_dir / "review" / (safe + "_review.png")).write_bytes(
                            review_png
                        )

                    done += 1
                    elapsed = time.time() - self.started_at
                    rate = done / elapsed if elapsed > 0 else 0
                    eta = (len(self.files) - done) / rate if rate > 0 else 0
                    self.progress_queue.put({
                        "type": "progress",
                        "done": done,
                        "total": len(self.files),
                        "current_file": fname,
                        "current_result": res,
                        "rate": rate,
                        "eta_sec": eta,
                    })

            if self.cancel_event.is_set():
                self.progress_queue.put({
                    "type": "cancelled",
                    "done": done,
                    "total": len(self.files),
                })
                return

            # Write output
            self._write_output()

            self.progress_queue.put({
                "type": "done",
                "results": self.results,
                "elapsed_sec": time.time() - self.started_at,
            })
        except Exception as e:
            self.progress_queue.put({
                "type": "error",
                "message": str(e),
                "traceback": traceback.format_exc(),
            })

    # -------------------------------------------------------------------
    # Output writers
    # -------------------------------------------------------------------

    def _max_questions(self) -> int:
        if not self.results:
            return 50
        seen = {r.sheet_type for r, _ in self.results}
        return max(TEMPLATES[s].n_questions for s in seen if s in TEMPLATES)

    def _row_for(self, serial: int, res: OmrResult, fname: str,
                 n_q: int) -> List:
        row = [serial, res.roll_number, res.set_letter]
        ans = list(res.answers) + [""] * (n_q - len(res.answers))
        row.extend(ans[:n_q])
        row.append(round(res.confidence * 100, 1))
        row.append("YES" if res.needs_review else "no")
        row.append(", ".join(res.review_items) if res.review_items else "")
        row.append(fname)
        if res.error:
            row.append(res.error)
        else:
            row.append("")
        return row

    def _header_row(self, n_q: int) -> List[str]:
        return (
            ["serial", "roll_number", "set"]
            + [f"Q{i+1}" for i in range(n_q)]
            + ["confidence", "needs_review", "review_items", "source_file", "error"]
        )

    def _write_output(self) -> None:
        n_q = self._max_questions()
        rows = [
            self._row_for(i + 1, res, fname, n_q)
            for i, (res, fname) in enumerate(self.results)
        ]
        header = self._header_row(n_q)

        if self.output_format == "csv":
            out_path = self.output_dir / "omr_results.csv"
            with out_path.open("w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(header)
                w.writerows(rows)

        elif self.output_format == "json":
            out_path = self.output_dir / "omr_results.json"
            data = []
            for i, (res, fname) in enumerate(self.results):
                d = {
                    "serial": i + 1,
                    "roll_number": res.roll_number,
                    "set": res.set_letter,
                    "answers": list(res.answers),
                    "confidence": round(res.confidence * 100, 1),
                    "needs_review": res.needs_review,
                    "review_items": res.review_items,
                    "source_file": fname,
                }
                if res.error:
                    d["error"] = res.error
                data.append(d)
            out_path.write_text(json.dumps(data, indent=2))

        else:  # xlsx
            try:
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font
            except ImportError:
                # Fall back to CSV if openpyxl isn't installed
                self.output_format = "csv"
                self._write_output()
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "OMR Results"
            ws.append(header)
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font
            review_fill = PatternFill("solid", fgColor="FFFFCC99")
            error_fill = PatternFill("solid", fgColor="FFFF9999")
            for row in rows:
                ws.append(row)
                last = ws.max_row
                # row[-2] = source_file, row[-3] = review_items, row[-4] = needs_review
                if row[-1]:  # error column
                    for cell in ws[last]:
                        cell.fill = error_fill
                elif row[-4] == "YES":
                    for cell in ws[last]:
                        cell.fill = review_fill
            # Auto-size first columns
            for col_letter in ["A", "B", "C"]:
                ws.column_dimensions[col_letter].width = 14
            wb.save(self.output_dir / "omr_results.xlsx")

        # Always also write a plain-text SUMMARY
        summary = []
        ok = sum(1 for r, _ in self.results if not r.error)
        rv = sum(1 for r, _ in self.results if r.needs_review)
        avg = (
            sum(r.confidence for r, _ in self.results) / len(self.results) * 100
            if self.results else 0
        )
        summary.append(f"Sheets scanned:        {len(self.results)}")
        summary.append(f"  OK:                  {ok}")
        summary.append(f"  Failed:              {len(self.results) - ok}")
        summary.append(f"  Needing review:      {rv}")
        summary.append(f"Average confidence:    {avg:.1f}%")
        summary.append(f"Elapsed:               "
                       f"{time.time() - self.started_at:.1f} seconds")
        summary.append("")
        summary.append("Per-sheet summary:")
        for i, (res, fname) in enumerate(self.results, start=1):
            line = (
                f"  {i:4d}. {fname[:45]:<45} "
                f"roll={res.roll_number} "
                f"set={res.set_letter}  "
                f"conf={res.confidence * 100:5.1f}%  "
                f"review={'YES' if res.needs_review else 'no'}"
            )
            if res.error:
                line += f"  ERROR: {res.error}"
            summary.append(line)
        (self.output_dir / "SUMMARY.txt").write_text("\n".join(summary))


# -----------------------------------------------------------------------------
# GUI
# -----------------------------------------------------------------------------


class OmrApp:
    """Tkinter GUI tying everything together."""

    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        root.title("OMR Scanner — Desktop")
        root.geometry("780x600")
        root.minsize(700, 500)

        self.input_dir = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.sheet_type = tk.StringVar(value="auto")
        self.output_format = tk.StringVar(value="xlsx")
        self.save_review = tk.BooleanVar(value=True)
        self.recursive = tk.BooleanVar(value=False)

        self.progress_queue: "queue.Queue[dict]" = queue.Queue()
        self.cancel_event = threading.Event()
        self.worker_thread: Optional[threading.Thread] = None

        self._build_ui()
        self._poll_queue()

    def _build_ui(self) -> None:
        pad = {"padx": 8, "pady": 4}

        frm = ttk.Frame(self.root, padding=12)
        frm.pack(fill=tk.BOTH, expand=True)

        # --- Inputs ----------------------------------------------------------
        title = ttk.Label(
            frm, text="OMR Scanner",
            font=("Segoe UI", 16, "bold"),
        )
        title.grid(row=0, column=0, columnspan=3, sticky=tk.W, **pad)

        ttk.Label(frm, text="Input folder:").grid(row=1, column=0, sticky=tk.E, **pad)
        ttk.Entry(frm, textvariable=self.input_dir, width=60).grid(
            row=1, column=1, sticky=tk.EW, **pad
        )
        ttk.Button(frm, text="Browse…", command=self._pick_input).grid(
            row=1, column=2, **pad
        )

        ttk.Label(frm, text="Output folder:").grid(row=2, column=0, sticky=tk.E, **pad)
        ttk.Entry(frm, textvariable=self.output_dir, width=60).grid(
            row=2, column=1, sticky=tk.EW, **pad
        )
        ttk.Button(frm, text="Browse…", command=self._pick_output).grid(
            row=2, column=2, **pad
        )

        # --- Options ---------------------------------------------------------
        opt = ttk.LabelFrame(frm, text="Options", padding=8)
        opt.grid(row=3, column=0, columnspan=3, sticky=tk.EW, **pad)

        ttk.Label(opt, text="Sheet type:").grid(row=0, column=0, sticky=tk.W, padx=4, pady=4)
        for i, (val, label) in enumerate([
            ("auto", "Auto-detect"),
            ("omr_50", "50-mark"),
            ("omr_100", "100-mark"),
        ]):
            ttk.Radiobutton(opt, text=label, value=val,
                            variable=self.sheet_type).grid(
                row=0, column=i + 1, sticky=tk.W, padx=8
            )

        ttk.Label(opt, text="Output format:").grid(row=1, column=0, sticky=tk.W, padx=4, pady=4)
        for i, (val, label) in enumerate([
            ("xlsx", "Excel (.xlsx)"),
            ("csv", "CSV"),
            ("json", "JSON"),
        ]):
            ttk.Radiobutton(opt, text=label, value=val,
                            variable=self.output_format).grid(
                row=1, column=i + 1, sticky=tk.W, padx=8
            )

        ttk.Checkbutton(
            opt, text="Save annotated review images",
            variable=self.save_review,
        ).grid(row=2, column=0, columnspan=3, sticky=tk.W, padx=4, pady=4)

        ttk.Checkbutton(
            opt, text="Search subfolders recursively",
            variable=self.recursive,
        ).grid(row=2, column=3, sticky=tk.W, padx=4, pady=4)

        # --- Actions ---------------------------------------------------------
        action_row = ttk.Frame(frm)
        action_row.grid(row=4, column=0, columnspan=3, sticky=tk.EW, **pad)
        self.start_btn = ttk.Button(
            action_row, text="Start Scanning",
            command=self._start,
        )
        self.start_btn.pack(side=tk.LEFT, padx=4)
        self.stop_btn = ttk.Button(
            action_row, text="Stop", command=self._stop, state=tk.DISABLED,
        )
        self.stop_btn.pack(side=tk.LEFT, padx=4)
        ttk.Button(
            action_row, text="Open Output Folder",
            command=self._open_output,
        ).pack(side=tk.LEFT, padx=4)

        # --- Progress --------------------------------------------------------
        prog_frm = ttk.LabelFrame(frm, text="Progress", padding=8)
        prog_frm.grid(row=5, column=0, columnspan=3, sticky=tk.NSEW, **pad)
        frm.rowconfigure(5, weight=1)
        frm.columnconfigure(1, weight=1)

        self.progress_var = tk.DoubleVar(value=0)
        self.progress_bar = ttk.Progressbar(
            prog_frm, mode="determinate", variable=self.progress_var,
        )
        self.progress_bar.pack(fill=tk.X, padx=4, pady=4)

        self.status_label = ttk.Label(
            prog_frm, text="Ready.", font=("Segoe UI", 10),
        )
        self.status_label.pack(anchor=tk.W, padx=4)

        log_frame = ttk.Frame(prog_frm)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        self.log = tk.Text(log_frame, height=12, wrap=tk.NONE)
        self.log.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scroll = ttk.Scrollbar(log_frame, command=self.log.yview)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.log.config(yscrollcommand=log_scroll.set)
        self.log.tag_config("ok", foreground="darkgreen")
        self.log.tag_config("review", foreground="darkorange")
        self.log.tag_config("error", foreground="red")

    # -------------------------------------------------------------------
    # Button handlers
    # -------------------------------------------------------------------

    def _pick_input(self) -> None:
        path = filedialog.askdirectory(title="Select folder of OMR sheets")
        if path:
            self.input_dir.set(path)
            if not self.output_dir.get():
                # Default output → input/<timestamp>
                ts = time.strftime("%Y%m%d_%H%M%S")
                self.output_dir.set(str(Path(path) / f"results_{ts}"))

    def _pick_output(self) -> None:
        path = filedialog.askdirectory(title="Select output folder")
        if path:
            self.output_dir.set(path)

    def _open_output(self) -> None:
        path = self.output_dir.get()
        if not path or not os.path.exists(path):
            messagebox.showinfo("Output", "No output folder to open yet.")
            return
        if sys.platform.startswith("win"):
            os.startfile(path)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            os.system(f"open '{path}'")
        else:
            os.system(f"xdg-open '{path}'")

    def _gather_files(self) -> List[Path]:
        inp = Path(self.input_dir.get())
        if not inp.is_dir():
            return []
        if self.recursive.get():
            files = [p for p in inp.rglob("*")
                     if p.is_file() and p.suffix.lower() in ScanJob.SUPPORTED_EXTS]
        else:
            files = [p for p in inp.iterdir()
                     if p.is_file() and p.suffix.lower() in ScanJob.SUPPORTED_EXTS]
        return sorted(files)

    def _start(self) -> None:
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showwarning("Already running", "A scan is already in progress.")
            return

        inp = self.input_dir.get().strip()
        out = self.output_dir.get().strip()
        if not inp or not os.path.isdir(inp):
            messagebox.showerror("Input", "Please choose a valid input folder.")
            return
        if not out:
            messagebox.showerror("Output", "Please choose an output folder.")
            return

        files = self._gather_files()
        if not files:
            messagebox.showerror(
                "No files",
                "Couldn't find any supported image files in the input folder. "
                f"Looked for: {', '.join(ScanJob.SUPPORTED_EXTS)}",
            )
            return

        self.log.delete("1.0", tk.END)
        self._log(f"Found {len(files)} file(s). Starting scan…")

        self.cancel_event.clear()
        self.progress_var.set(0)
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)

        job = ScanJob(
            files=files,
            sheet_type=self.sheet_type.get(),
            output_dir=Path(out),
            output_format=self.output_format.get(),
            save_review_images=self.save_review.get(),
            progress_queue=self.progress_queue,
            cancel_event=self.cancel_event,
        )
        self.worker_thread = threading.Thread(target=job.run, daemon=True)
        self.worker_thread.start()

    def _stop(self) -> None:
        if messagebox.askyesno("Stop", "Stop the current scan?"):
            self.cancel_event.set()
            self._log("Cancelling…", "error")

    def _log(self, text: str, tag: str = "") -> None:
        if tag:
            self.log.insert(tk.END, text + "\n", tag)
        else:
            self.log.insert(tk.END, text + "\n")
        self.log.see(tk.END)

    # -------------------------------------------------------------------
    # Queue poller — pulls progress events from the worker thread
    # -------------------------------------------------------------------

    def _poll_queue(self) -> None:
        try:
            while True:
                msg = self.progress_queue.get_nowait()
                self._handle_msg(msg)
        except queue.Empty:
            pass
        self.root.after(100, self._poll_queue)

    def _handle_msg(self, msg: dict) -> None:
        kind = msg.get("type")
        if kind == "start":
            self.status_label.config(text=f"Starting… {msg['total']} files")
        elif kind == "progress":
            pct = msg["done"] * 100 / msg["total"]
            self.progress_var.set(pct)
            eta_str = f"{int(msg['eta_sec'])}s" if msg["eta_sec"] > 1 else "0s"
            self.status_label.config(
                text=(
                    f"{msg['done']}/{msg['total']} ({pct:.0f}%)  "
                    f"{msg['rate']:.1f}/sec  ETA {eta_str}"
                )
            )
            res = msg["current_result"]
            tag = "ok"
            line = (
                f"  {msg['done']:4d}. {msg['current_file'][:38]:<38} "
                f"roll={res.roll_number}  set={res.set_letter}  "
                f"conf={res.confidence * 100:5.1f}%"
            )
            if res.error:
                tag = "error"
                line += f"  ERROR: {res.error[:60]}"
            elif res.needs_review:
                tag = "review"
                line += "  REVIEW"
            self._log(line, tag)
        elif kind == "done":
            self.progress_var.set(100)
            self.status_label.config(
                text=f"Done in {msg['elapsed_sec']:.1f}s. "
                     f"Output saved to {self.output_dir.get()}"
            )
            self._log(
                f"\n✓ Finished {len(msg['results'])} sheets in "
                f"{msg['elapsed_sec']:.1f}s",
                "ok",
            )
            self.start_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)
            messagebox.showinfo(
                "Scan complete",
                f"Processed {len(msg['results'])} sheets in "
                f"{msg['elapsed_sec']:.1f} seconds.\n\n"
                f"Output: {self.output_dir.get()}",
            )
        elif kind == "cancelled":
            self.status_label.config(text="Cancelled.")
            self._log("Scan cancelled.", "error")
            self.start_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)
        elif kind == "error":
            self.status_label.config(text="Error.")
            self._log(f"ERROR: {msg['message']}", "error")
            self._log(msg.get("traceback", ""), "error")
            self.start_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)


def main() -> None:
    root = tk.Tk()
    # Use a nicer theme where available
    try:
        style = ttk.Style()
        if "clam" in style.theme_names():
            style.theme_use("clam")
    except Exception:
        pass
    OmrApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
