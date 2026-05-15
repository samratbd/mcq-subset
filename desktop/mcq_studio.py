#!/usr/bin/env python3
"""MCQ Studio — desktop application bundling OMR Scanner + MCQ Shuffler.

Two-tab Tkinter app:

  1. OMR Scanner   — batch-process scanned answer sheets
  2. MCQ Shuffler  — generate Set A / B / C / ... variants from a question bank

Both run entirely on the user's machine. No web server, no internet, no
upload limits, no time-outs. State (last-used folders, options) is
saved between runs to ~/.mcq_studio/config.json.

Run with:
    python mcq_studio.py

Build into a standalone Windows .exe:
    pip install pyinstaller
    pyinstaller --noconfirm mcq_studio.spec

The .spec file ships next to this script.
"""

from __future__ import annotations

import csv
import io
import json
import os
import queue
import re
import shutil
import subprocess
import sys
import threading
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Callable, List, Optional, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# Locate the bundled `app/` package — both when running from source and
# when frozen by PyInstaller into a one-file exe.
# When frozen by PyInstaller, sys._MEIPASS contains everything.
# When running from source, we need the PROJECT ROOT (one level up from desktop/)
# so that `app/` package is importable.
if getattr(sys, "frozen", False):
    _BASE_DIR = Path(sys._MEIPASS)
else:
    _BASE_DIR = Path(__file__).resolve().parent.parent  # project root
sys.path.insert(0, str(_BASE_DIR))

# Import the same code the web server uses.
try:
    from app.omr import scan_omr, scan_and_render, render_review_image, TEMPLATES
    from app.omr.scanner import OmrResult
    from app.parsers import parse_upload
    from app.shuffler import make_set
    from app.writers.docx_writer import write_docx_normal, write_docx_database
    from app.writers.xlsx_writer import write_xlsx
    from app.writers.csv_writer import write_csv
    from app.writers.pdf_writer import (
        docx_bytes_to_pdf_bytes, pdf_engine_available, pdf_engine_name
    )
except ImportError as e:
    # Without the app package we can't run. Show a clear error.
    err_msg = (
        "MCQ Studio cannot start — required modules are missing.\n\n"
        f"Details: {e}\n\n"
        "If you're running from source, install dependencies:\n"
        "  pip install -r requirements.txt\n\n"
        "If this is the installed app, please reinstall."
    )
    try:
        # GUI message if possible
        _root = tk.Tk()
        _root.withdraw()
        messagebox.showerror("MCQ Studio — Startup error", err_msg)
    except Exception:
        print(err_msg, file=sys.stderr)
    sys.exit(1)


APP_NAME = "Onurion OMR Studio"
APP_VERSION = "1.0"
APP_DEVELOPER = "S M Samrat"
APP_COMPANY = "Onurion.com"
APP_COPYRIGHT = f"Developer: {APP_DEVELOPER}  |  Company: {APP_COMPANY}"

CONFIG_DIR = Path.home() / ".mcq_studio"
CONFIG_FILE = CONFIG_DIR / "config.json"


def _load_default_header() -> Optional[bytes]:
    """Return the default header image bytes, or None if not found.

    Looks for the bundled default_header.jpg in the same location as the
    web app uses:  static/assets/default_header.jpg
    Works both when running from source and when frozen by PyInstaller.
    """
    candidates = [
        _BASE_DIR / "static" / "assets" / "default_header.jpg",
        _BASE_DIR / "app" / "default_header.jpg",
        _BASE_DIR / "desktop" / "static" / "assets" / "default_header.jpg",
        Path(__file__).parent / "static" / "assets" / "default_header.jpg",
        Path(__file__).parent.parent / "static" / "assets" / "default_header.jpg",
    ]
    for p in candidates:
        if p.exists():
            try:
                return p.read_bytes()
            except Exception:
                pass
    return None


# ============================================================================
# Persistent settings
# ============================================================================


class Settings:
    """Load and save user preferences across sessions."""

    DEFAULTS = {
        "omr_input_dir": str(Path.home() / "Documents"),
        "omr_output_dir": str(Path.home() / "Documents" / "MCQ Studio" / "OMR Results"),
        "omr_sheet_type": "auto",
        "omr_output_format": "xlsx",
        "omr_save_review": True,
        "omr_recursive": False,

        "shuffler_input_file": "",
        "shuffler_output_dir": str(Path.home() / "Documents" / "MCQ Studio" / "Shuffled Sets"),
        "shuffler_n_sets": 4,
        "shuffler_layout": "normal",
        "shuffler_output_format": "docx",
        "shuffler_shuffle_options": True,
    }

    def __init__(self) -> None:
        self.data = dict(self.DEFAULTS)
        self.load()

    def load(self) -> None:
        if CONFIG_FILE.exists():
            try:
                loaded = json.loads(CONFIG_FILE.read_text())
                self.data.update({k: v for k, v in loaded.items()
                                  if k in self.DEFAULTS})
            except Exception:
                pass  # keep defaults

    def save(self) -> None:
        try:
            CONFIG_DIR.mkdir(parents=True, exist_ok=True)
            CONFIG_FILE.write_text(json.dumps(self.data, indent=2))
        except Exception:
            pass  # not critical

    def get(self, key: str):
        return self.data.get(key, self.DEFAULTS.get(key))

    def set(self, key: str, value) -> None:
        self.data[key] = value


# ============================================================================
# OMR scan job (background worker)
# ============================================================================


class ScanJob:
    """One run of the OMR scanner over a list of files."""

    SUPPORTED_EXTS = (".bmp", ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".gif")

    def __init__(
        self,
        files: List[Path],
        sheet_type: str,
        output_dir: Path,
        output_format: str,
        save_review_images: bool,
        progress_queue: "queue.Queue[dict]",
        cancel_event: threading.Event,
    ) -> None:
        self.files = files
        self.sheet_type = sheet_type
        self.output_dir = output_dir
        self.output_format = output_format
        self.save_review_images = save_review_images
        self.progress_queue = progress_queue
        self.cancel_event = cancel_event
        self.results: List[Tuple[OmrResult, str]] = []
        self.started_at = time.time()

    def _scan_one(self, file_path: Path) -> Tuple[OmrResult, str, Optional[bytes]]:
        if self.cancel_event.is_set():
            return (self._make_error_result("Cancelled"), file_path.name, None)
        try:
            data = file_path.read_bytes()
            if self.save_review_images:
                # 2× faster combined call — single warp for both scan + render
                res, review_png = scan_and_render(data, sheet_type=self.sheet_type)
            else:
                res = scan_omr(data, sheet_type=self.sheet_type)
                review_png = None
        except Exception as e:
            return (self._make_error_result(f"{type(e).__name__}: {e}"),
                    file_path.name, None)
        if res.error:
            review_png = None
        return res, file_path.name, review_png

    def _make_error_result(self, msg: str) -> OmrResult:
        return OmrResult(
            sheet_type=("omr_50" if self.sheet_type == "auto" else self.sheet_type),
            roll_number="?", set_letter="?", answers=[],
            confidence=0.0, needs_review=True,
            review_items=["scan_failed"], fill_fractions=[],
            error=msg,
        )

    def run(self) -> None:
        try:
            self.output_dir.mkdir(parents=True, exist_ok=True)
            if self.save_review_images:
                (self.output_dir / "review").mkdir(exist_ok=True)

            self.progress_queue.put({"type": "start", "total": len(self.files)})
            max_workers = min(8, max(2, (os.cpu_count() or 2)))

            done = 0
            with ThreadPoolExecutor(max_workers=max_workers) as pool:
                futures = {pool.submit(self._scan_one, fp): fp for fp in self.files}
                for fut in as_completed(futures):
                    if self.cancel_event.is_set():
                        for f in futures:
                            f.cancel()
                        break
                    fp = futures[fut]
                    try:
                        res, fname, review_png = fut.result()
                    except Exception as e:
                        res = self._make_error_result(str(e))
                        fname = fp.name
                        review_png = None

                    self.results.append((res, fname))

                    if review_png:
                        safe = re.sub(r"[^A-Za-z0-9._-]+", "_", fname)
                        (self.output_dir / "review" / (safe + "_review.png")).write_bytes(
                            review_png
                        )

                    done += 1
                    elapsed = time.time() - self.started_at
                    rate = done / elapsed if elapsed > 0 else 0
                    eta = (len(self.files) - done) / rate if rate > 0 else 0
                    self.progress_queue.put({
                        "type": "progress",
                        "done": done, "total": len(self.files),
                        "current_file": fname,
                        "current_result": res,
                        "rate": rate, "eta_sec": eta,
                    })

            if self.cancel_event.is_set():
                self.progress_queue.put({"type": "cancelled", "done": done,
                                         "total": len(self.files)})
                return

            self._write_output()
            self.progress_queue.put({
                "type": "done",
                "results": self.results,
                "elapsed_sec": time.time() - self.started_at,
            })
        except Exception as e:
            self.progress_queue.put({
                "type": "error",
                "message": str(e),
                "traceback": traceback.format_exc(),
            })

    def _max_questions(self) -> int:
        if not self.results:
            return 50
        seen = {r.sheet_type for r, _ in self.results}
        return max(TEMPLATES[s].n_questions for s in seen if s in TEMPLATES)

    def _row_for(self, serial, res, fname, n_q):
        row = [serial, res.roll_number, res.set_letter]
        ans = list(res.answers) + [""] * (n_q - len(res.answers))
        row.extend(ans[:n_q])
        row.append(round(res.confidence * 100, 1))
        row.append("YES" if res.needs_review else "no")
        row.append(", ".join(res.review_items) if res.review_items else "")
        row.append(fname)
        row.append(res.error or "")
        return row

    def _header_row(self, n_q):
        return (
            ["serial", "roll_number", "set"]
            + [f"Q{i+1}" for i in range(n_q)]
            + ["confidence", "needs_review", "review_items",
               "source_file", "error"]
        )

    def _write_output(self) -> None:
        n_q = self._max_questions()
        rows = [self._row_for(i + 1, res, fname, n_q)
                for i, (res, fname) in enumerate(self.results)]
        header = self._header_row(n_q)

        if self.output_format == "csv":
            out = self.output_dir / "omr_results.csv"
            with out.open("w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(header)
                w.writerows(rows)
        elif self.output_format == "json":
            out = self.output_dir / "omr_results.json"
            data = []
            for i, (res, fname) in enumerate(self.results):
                d = {"serial": i + 1, "roll_number": res.roll_number,
                     "set": res.set_letter, "answers": list(res.answers),
                     "confidence": round(res.confidence * 100, 1),
                     "needs_review": res.needs_review,
                     "review_items": res.review_items,
                     "source_file": fname}
                if res.error:
                    d["error"] = res.error
                data.append(d)
            out.write_text(json.dumps(data, indent=2))
        else:  # xlsx
            try:
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font
            except ImportError:
                self.output_format = "csv"
                self._write_output()
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "OMR Results"
            ws.append(header)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            review_fill = PatternFill("solid", fgColor="FFFFCC99")
            error_fill = PatternFill("solid", fgColor="FFFF9999")
            for row in rows:
                ws.append(row)
                last = ws.max_row
                if row[-1]:
                    for cell in ws[last]:
                        cell.fill = error_fill
                elif row[-4] == "YES":
                    for cell in ws[last]:
                        cell.fill = review_fill
            for c in ["A", "B", "C"]:
                ws.column_dimensions[c].width = 14
            wb.save(self.output_dir / "omr_results.xlsx")

        # Plain-text summary
        ok = sum(1 for r, _ in self.results if not r.error)
        rv = sum(1 for r, _ in self.results if r.needs_review)
        avg = (sum(r.confidence for r, _ in self.results) / len(self.results)
               * 100 if self.results else 0)
        lines = [
            f"{APP_NAME} — OMR results",
            "=" * 40,
            f"Sheets scanned:        {len(self.results)}",
            f"  OK:                  {ok}",
            f"  Failed:              {len(self.results) - ok}",
            f"  Needing review:      {rv}",
            f"Average confidence:    {avg:.1f}%",
            f"Elapsed:               "
            f"{time.time() - self.started_at:.1f} seconds",
            "",
            "Per-sheet summary:",
        ]
        for i, (res, fname) in enumerate(self.results, start=1):
            line = (f"  {i:4d}. {fname[:45]:<45} "
                    f"roll={res.roll_number} set={res.set_letter}  "
                    f"conf={res.confidence * 100:5.1f}%  "
                    f"review={'YES' if res.needs_review else 'no'}")
            if res.error:
                line += f"  ERROR: {res.error}"
            lines.append(line)
        (self.output_dir / "SUMMARY.txt").write_text("\n".join(lines))


# ============================================================================
# Shuffler job (background worker)
# ============================================================================


class ShuffleJob:
    """Generate N sets from an input question bank."""

    def __init__(
        self,
        input_file: Path,
        output_dir: Path,
        n_sets: int,
        layout: str,
        output_format: str,
        shuffle_options: bool,
        title_prefix: str,
        progress_queue: "queue.Queue[dict]",
        cancel_event: threading.Event,
    ) -> None:
        self.input_file = input_file
        self.output_dir = output_dir
        self.n_sets = n_sets
        self.layout = layout
        self.output_format = output_format
        self.shuffle_options = shuffle_options
        self.title_prefix = title_prefix
        self.progress_queue = progress_queue
        self.cancel_event = cancel_event
        self.started_at = time.time()

    def run(self) -> None:
        try:
            self.output_dir.mkdir(parents=True, exist_ok=True)
            self.progress_queue.put({"type": "start", "total": self.n_sets})

            # Parse input
            try:
                questions = parse_upload(
                    self.input_file.name,
                    self.input_file.read_bytes(),
                )
            except Exception as e:
                self.progress_queue.put({
                    "type": "error",
                    "message": f"Cannot read input: {type(e).__name__}: {e}",
                    "traceback": traceback.format_exc(),
                })
                return

            if not questions:
                self.progress_queue.put({
                    "type": "error",
                    "message": "Input file has no questions to shuffle.",
                })
                return

            self.progress_queue.put({
                "type": "log",
                "text": f"Loaded {len(questions)} questions from "
                        f"{self.input_file.name}",
            })

            for set_idx in range(self.n_sets):
                if self.cancel_event.is_set():
                    break
                set_letter = chr(ord("A") + set_idx)
                title = f"{self.title_prefix} — Set {set_letter}".strip(" —")

                shuffled = make_set(
                    questions,
                    set_letter.lower(),
                    set_idx + 1,
                    shuffle_questions=True,
                    shuffle_options=self.shuffle_options,
                )

                try:
                    self._write_set(set_letter, title, shuffled)
                except Exception as e:
                    self.progress_queue.put({
                        "type": "log",
                        "text": f"ERROR writing Set {set_letter}: {e}",
                    })
                    continue

                self.progress_queue.put({
                    "type": "progress",
                    "done": set_idx + 1,
                    "total": self.n_sets,
                    "current": f"Set {set_letter}",
                })

            if self.cancel_event.is_set():
                self.progress_queue.put({"type": "cancelled"})
                return

            self.progress_queue.put({
                "type": "done",
                "elapsed_sec": time.time() - self.started_at,
            })
        except Exception as e:
            self.progress_queue.put({
                "type": "error",
                "message": str(e),
                "traceback": traceback.format_exc(),
            })

    def _write_set(self, set_letter: str, title: str, shuffled) -> None:
        fmt = self.output_format
        out_dir = self.output_dir

        # Load the default header image bundled with the app
        header_image = _load_default_header()

        if fmt in ("docx", "pdf"):
            if self.layout == "database":
                data = write_docx_database(
                    shuffled, title=title,
                    header_image=header_image,
                )
            else:
                data = write_docx_normal(
                    shuffled, title=title,
                    math_mode="equation",
                    header_image=header_image,
                )
            docx_path = out_dir / f"Set_{set_letter}.docx"
            docx_path.write_bytes(data)

            if fmt == "pdf":
                if not pdf_engine_available():
                    self.progress_queue.put({
                        "type": "log",
                        "text": (
                            f"PDF needs Microsoft Word OR LibreOffice installed.\n"
                            f"    → Kept Set_{set_letter}.docx instead.\n"
                            f"    → To enable PDF: pip install docx2pdf\n"
                            f"      (requires Microsoft Word on this PC)"
                        ),
                    })
                else:
                    try:
                        pdf_data = docx_bytes_to_pdf_bytes(data)
                        (out_dir / f"Set_{set_letter}.pdf").write_bytes(pdf_data)
                        # Remove the .docx if PDF succeeded
                        docx_path.unlink(missing_ok=True)
                    except Exception as e:
                        self.progress_queue.put({
                            "type": "log",
                            "text": f"PDF failed for Set {set_letter} (kept .docx): {e}",
                        })
        elif fmt == "xlsx":
            data = write_xlsx(shuffled, math_mode="katex")
            (out_dir / f"Set_{set_letter}.xlsx").write_bytes(data)
        elif fmt == "csv":
            data = write_csv(shuffled, math_mode="katex")
            (out_dir / f"Set_{set_letter}.csv").write_bytes(data)
        else:
            raise ValueError(f"Unknown output_format: {fmt}")


# ============================================================================
# Helpers
# ============================================================================


def open_folder(path: str) -> None:
    """Open a folder in the OS file explorer."""
    if not path or not os.path.exists(path):
        return
    if sys.platform.startswith("win"):
        os.startfile(path)  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        subprocess.Popen(["open", path])
    else:
        subprocess.Popen(["xdg-open", path])


def ensure_dir(path: str) -> None:
    """Create the directory if it doesn't exist."""
    try:
        Path(path).mkdir(parents=True, exist_ok=True)
    except Exception:
        pass


# ============================================================================
# GUI — OMR tab
# ============================================================================


class OmrTab(ttk.Frame):
    def __init__(self, parent: ttk.Notebook, settings: Settings) -> None:
        super().__init__(parent, padding=10)
        self.settings = settings

        self.input_dir = tk.StringVar(value=settings.get("omr_input_dir"))
        self.output_dir = tk.StringVar(value=settings.get("omr_output_dir"))
        self.sheet_type = tk.StringVar(value=settings.get("omr_sheet_type"))
        self.output_format = tk.StringVar(value=settings.get("omr_output_format"))
        self.save_review = tk.BooleanVar(value=settings.get("omr_save_review"))
        self.recursive = tk.BooleanVar(value=settings.get("omr_recursive"))

        self.progress_queue: "queue.Queue[dict]" = queue.Queue()
        self.cancel_event = threading.Event()
        self.worker_thread: Optional[threading.Thread] = None

        self._build_ui()
        self.after(100, self._poll_queue)

    def _build_ui(self) -> None:
        pad = {"padx": 6, "pady": 4}

        title = ttk.Label(self, text="OMR Sheet Scanner",
                          font=("Segoe UI", 14, "bold"))
        title.grid(row=0, column=0, columnspan=3, sticky=tk.W, **pad)

        ttk.Label(self, text="Input folder:").grid(
            row=1, column=0, sticky=tk.E, **pad)
        ttk.Entry(self, textvariable=self.input_dir, width=55).grid(
            row=1, column=1, sticky=tk.EW, **pad)
        ttk.Button(self, text="Browse…", command=self._pick_input).grid(
            row=1, column=2, **pad)

        ttk.Label(self, text="Output folder:").grid(
            row=2, column=0, sticky=tk.E, **pad)
        ttk.Entry(self, textvariable=self.output_dir, width=55).grid(
            row=2, column=1, sticky=tk.EW, **pad)
        ttk.Button(self, text="Browse…", command=self._pick_output).grid(
            row=2, column=2, **pad)

        opt = ttk.LabelFrame(self, text="Options", padding=8)
        opt.grid(row=3, column=0, columnspan=3, sticky=tk.EW, **pad)

        ttk.Label(opt, text="Sheet type:").grid(row=0, column=0,
                                                 sticky=tk.W, padx=4)
        for i, (val, lbl) in enumerate([
            ("auto", "Auto-detect"),
            ("omr_50", "50-mark"),
            ("omr_100", "100-mark"),
        ]):
            ttk.Radiobutton(opt, text=lbl, value=val,
                            variable=self.sheet_type).grid(
                row=0, column=i + 1, sticky=tk.W, padx=8)

        ttk.Label(opt, text="Output format:").grid(row=1, column=0,
                                                    sticky=tk.W, padx=4, pady=4)
        for i, (val, lbl) in enumerate([
            ("xlsx", "Excel (.xlsx)"),
            ("csv", "CSV"),
            ("json", "JSON"),
        ]):
            ttk.Radiobutton(opt, text=lbl, value=val,
                            variable=self.output_format).grid(
                row=1, column=i + 1, sticky=tk.W, padx=8)

        ttk.Checkbutton(opt, text="Save annotated review images",
                        variable=self.save_review).grid(
            row=2, column=0, columnspan=2, sticky=tk.W, padx=4, pady=4)
        ttk.Checkbutton(opt, text="Search subfolders recursively",
                        variable=self.recursive).grid(
            row=2, column=2, columnspan=2, sticky=tk.W, padx=4, pady=4)

        # Actions
        act = ttk.Frame(self)
        act.grid(row=4, column=0, columnspan=3, sticky=tk.EW, **pad)
        self.start_btn = ttk.Button(act, text="▶ Start Scanning",
                                     command=self._start)
        self.start_btn.pack(side=tk.LEFT, padx=4)
        self.stop_btn = ttk.Button(act, text="■ Stop",
                                    command=self._stop, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=4)
        ttk.Button(act, text="🔄 Clear",
                   command=self._clear).pack(side=tk.LEFT, padx=4)
        ttk.Button(act, text="📂 Open Output",
                   command=lambda: open_folder(self.output_dir.get())
                   ).pack(side=tk.LEFT, padx=4)

        # Progress
        prog = ttk.LabelFrame(self, text="Progress", padding=8)
        prog.grid(row=5, column=0, columnspan=3, sticky=tk.NSEW, **pad)
        self.rowconfigure(5, weight=1)
        self.columnconfigure(1, weight=1)

        self.progress_var = tk.DoubleVar(value=0)
        ttk.Progressbar(prog, mode="determinate",
                        variable=self.progress_var).pack(fill=tk.X, padx=4, pady=4)
        self.status_label = ttk.Label(prog, text="Ready.")
        self.status_label.pack(anchor=tk.W, padx=4)

        log_frame = ttk.Frame(prog)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        self.log = tk.Text(log_frame, height=12, wrap=tk.NONE,
                            font=("Consolas", 9))
        self.log.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb = ttk.Scrollbar(log_frame, command=self.log.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.log.config(yscrollcommand=sb.set)
        self.log.tag_config("ok", foreground="#1a7e2c")
        self.log.tag_config("review", foreground="#cc7000")
        self.log.tag_config("error", foreground="#cc0000")

    def _pick_input(self):
        path = filedialog.askdirectory(
            title="Select folder of OMR sheets",
            initialdir=self.input_dir.get() or str(Path.home()),
        )
        if path:
            self.input_dir.set(path)
            self.settings.set("omr_input_dir", path)
            self.settings.save()

    def _pick_output(self):
        path = filedialog.askdirectory(
            title="Select output folder",
            initialdir=self.output_dir.get() or str(Path.home()),
        )
        if path:
            self.output_dir.set(path)
            self.settings.set("omr_output_dir", path)
            self.settings.save()

    def _gather_files(self) -> List[Path]:
        inp = Path(self.input_dir.get())
        if not inp.is_dir():
            return []
        if self.recursive.get():
            files = [p for p in inp.rglob("*")
                     if p.is_file() and p.suffix.lower() in ScanJob.SUPPORTED_EXTS]
        else:
            files = [p for p in inp.iterdir()
                     if p.is_file() and p.suffix.lower() in ScanJob.SUPPORTED_EXTS]
        return sorted(files)

    def _start(self):
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showwarning("Already running",
                                    "A scan is already in progress.")
            return
        inp = self.input_dir.get().strip()
        out = self.output_dir.get().strip()
        if not inp or not os.path.isdir(inp):
            messagebox.showerror("Input",
                                  "Please choose a valid input folder.")
            return
        if not out:
            messagebox.showerror("Output",
                                  "Please choose an output folder.")
            return

        files = self._gather_files()
        if not files:
            messagebox.showerror(
                "No files",
                "No supported image files in input folder.\n"
                f"Looked for: {', '.join(ScanJob.SUPPORTED_EXTS)}")
            return

        # Save settings
        self.settings.set("omr_sheet_type", self.sheet_type.get())
        self.settings.set("omr_output_format", self.output_format.get())
        self.settings.set("omr_save_review", self.save_review.get())
        self.settings.set("omr_recursive", self.recursive.get())
        self.settings.save()

        self.log.delete("1.0", tk.END)
        self._log(f"Found {len(files)} file(s). Starting…")
        self.cancel_event.clear()
        self.progress_var.set(0)
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)

        job = ScanJob(
            files=files,
            sheet_type=self.sheet_type.get(),
            output_dir=Path(out),
            output_format=self.output_format.get(),
            save_review_images=self.save_review.get(),
            progress_queue=self.progress_queue,
            cancel_event=self.cancel_event,
        )
        self.worker_thread = threading.Thread(target=job.run, daemon=True)
        self.worker_thread.start()

    def _stop(self):
        if messagebox.askyesno("Stop", "Stop the current scan?"):
            self.cancel_event.set()
            self._log("Cancelling…", "error")

    def _clear(self):
        """Reset the UI for a fresh run — clears log and progress bar."""
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showwarning("Busy", "A scan is running. Stop it first.")
            return
        self.log.delete("1.0", tk.END)
        self.progress_var.set(0)
        self.status_label.config(text="Ready.")
        self._log("Cleared. Ready for a new scan.")

    def _log(self, text: str, tag: str = "") -> None:
        if tag:
            self.log.insert(tk.END, text + "\n", tag)
        else:
            self.log.insert(tk.END, text + "\n")
        self.log.see(tk.END)

    def _poll_queue(self):
        try:
            while True:
                msg = self.progress_queue.get_nowait()
                self._handle_msg(msg)
        except queue.Empty:
            pass
        self.after(100, self._poll_queue)

    def _handle_msg(self, msg):
        kind = msg.get("type")
        if kind == "start":
            self.status_label.config(text=f"Starting… {msg['total']} files")
        elif kind == "progress":
            pct = msg["done"] * 100 / msg["total"]
            self.progress_var.set(pct)
            eta_str = f"{int(msg['eta_sec'])}s" if msg["eta_sec"] > 1 else "0s"
            self.status_label.config(
                text=f"{msg['done']}/{msg['total']} ({pct:.0f}%)  "
                     f"{msg['rate']:.1f}/sec  ETA {eta_str}")
            res = msg["current_result"]
            tag = "ok"
            line = (f"  {msg['done']:4d}. {msg['current_file'][:36]:<36} "
                    f"roll={res.roll_number}  set={res.set_letter}  "
                    f"conf={res.confidence * 100:5.1f}%")
            if res.error:
                tag = "error"
                line += f"  ERROR: {res.error[:50]}"
            elif res.needs_review:
                tag = "review"
                line += "  REVIEW"
            self._log(line, tag)
        elif kind == "done":
            self.progress_var.set(100)
            self.status_label.config(
                text=f"Done in {msg['elapsed_sec']:.1f}s. "
                     f"Output: {self.output_dir.get()}")
            self._log(f"\n✓ Finished {len(msg['results'])} sheets in "
                       f"{msg['elapsed_sec']:.1f}s", "ok")
            self.start_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)
            ok = sum(1 for r, _ in msg["results"] if not r.error)
            rv = sum(1 for r, _ in msg["results"] if r.needs_review)
            messagebox.showinfo(
                "Scan complete",
                f"Processed {len(msg['results'])} sheets in "
                f"{msg['elapsed_sec']:.1f} seconds.\n\n"
                f"  Successful: {ok}\n"
                f"  Need review: {rv}\n\n"
                f"Output: {self.output_dir.get()}",
            )
        elif kind == "cancelled":
            self.status_label.config(text="Cancelled.")
            self._log("Scan cancelled.", "error")
            self.start_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)
        elif kind == "error":
            self.status_label.config(text="Error.")
            self._log(f"ERROR: {msg['message']}", "error")
            if msg.get("traceback"):
                self._log(msg["traceback"], "error")
            self.start_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)


# ============================================================================
# GUI — Shuffler tab
# ============================================================================


class ShufflerTab(ttk.Frame):
    SUPPORTED_INPUTS = (".docx", ".xlsx", ".csv")

    def __init__(self, parent: ttk.Notebook, settings: Settings) -> None:
        super().__init__(parent, padding=10)
        self.settings = settings

        self.input_file = tk.StringVar(value=settings.get("shuffler_input_file"))
        self.output_dir = tk.StringVar(value=settings.get("shuffler_output_dir"))
        self.n_sets = tk.IntVar(value=settings.get("shuffler_n_sets"))
        self.layout = tk.StringVar(value=settings.get("shuffler_layout"))
        self.output_format = tk.StringVar(value=settings.get("shuffler_output_format"))
        self.shuffle_options = tk.BooleanVar(value=settings.get("shuffler_shuffle_options"))
        self.title_prefix = tk.StringVar(value="Test")

        self.progress_queue: "queue.Queue[dict]" = queue.Queue()
        self.cancel_event = threading.Event()
        self.worker_thread: Optional[threading.Thread] = None

        self._build_ui()
        self.after(100, self._poll_queue)

    def _build_ui(self):
        pad = {"padx": 6, "pady": 4}
        title = ttk.Label(self, text="MCQ Shuffler — generate Set A, B, C…",
                          font=("Segoe UI", 14, "bold"))
        title.grid(row=0, column=0, columnspan=3, sticky=tk.W, **pad)

        ttk.Label(self, text="Question bank:").grid(
            row=1, column=0, sticky=tk.E, **pad)
        ttk.Entry(self, textvariable=self.input_file, width=55).grid(
            row=1, column=1, sticky=tk.EW, **pad)
        ttk.Button(self, text="Browse…",
                   command=self._pick_input).grid(row=1, column=2, **pad)

        ttk.Label(self, text="Output folder:").grid(
            row=2, column=0, sticky=tk.E, **pad)
        ttk.Entry(self, textvariable=self.output_dir, width=55).grid(
            row=2, column=1, sticky=tk.EW, **pad)
        ttk.Button(self, text="Browse…",
                   command=self._pick_output).grid(row=2, column=2, **pad)

        ttk.Label(self, text="Title prefix:").grid(
            row=3, column=0, sticky=tk.E, **pad)
        ttk.Entry(self, textvariable=self.title_prefix, width=55).grid(
            row=3, column=1, sticky=tk.EW, **pad)

        opt = ttk.LabelFrame(self, text="Options", padding=8)
        opt.grid(row=4, column=0, columnspan=3, sticky=tk.EW, **pad)

        ttk.Label(opt, text="Number of sets:").grid(row=0, column=0,
                                                     sticky=tk.W, padx=4)
        ttk.Spinbox(opt, from_=1, to=26, textvariable=self.n_sets,
                    width=5).grid(row=0, column=1, sticky=tk.W, padx=4)

        ttk.Label(opt, text="Layout:").grid(row=0, column=2,
                                             sticky=tk.W, padx=20)
        for i, (val, lbl) in enumerate([
            ("normal", "Normal (2-column)"),
            ("database", "Database (table)"),
        ]):
            ttk.Radiobutton(opt, text=lbl, value=val,
                            variable=self.layout).grid(
                row=0, column=3 + i, sticky=tk.W, padx=4)

        ttk.Label(opt, text="Output format:").grid(row=1, column=0,
                                                    sticky=tk.W, padx=4, pady=4)
        for i, (val, lbl) in enumerate([
            ("docx", "Word (.docx)"),
            ("pdf", "PDF"),
            ("xlsx", "Excel"),
            ("csv", "CSV"),
        ]):
            ttk.Radiobutton(opt, text=lbl, value=val,
                            variable=self.output_format).grid(
                row=1, column=i + 1, sticky=tk.W, padx=8)

        ttk.Checkbutton(opt, text="Shuffle option order within each question",
                        variable=self.shuffle_options).grid(
            row=2, column=0, columnspan=4, sticky=tk.W, padx=4, pady=4)

        # Actions
        act = ttk.Frame(self)
        act.grid(row=5, column=0, columnspan=3, sticky=tk.EW, **pad)
        self.start_btn = ttk.Button(act, text="▶ Generate Sets",
                                     command=self._start)
        self.start_btn.pack(side=tk.LEFT, padx=4)
        self.stop_btn = ttk.Button(act, text="■ Stop",
                                    command=self._stop, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=4)
        ttk.Button(act, text="🔄 Clear",
                   command=self._clear).pack(side=tk.LEFT, padx=4)
        ttk.Button(act, text="📂 Open Output",
                   command=lambda: open_folder(self.output_dir.get())
                   ).pack(side=tk.LEFT, padx=4)

        # Progress
        prog = ttk.LabelFrame(self, text="Progress", padding=8)
        prog.grid(row=6, column=0, columnspan=3, sticky=tk.NSEW, **pad)
        self.rowconfigure(6, weight=1)
        self.columnconfigure(1, weight=1)

        self.progress_var = tk.DoubleVar(value=0)
        ttk.Progressbar(prog, mode="determinate",
                        variable=self.progress_var).pack(fill=tk.X, padx=4, pady=4)
        self.status_label = ttk.Label(prog, text="Ready.")
        self.status_label.pack(anchor=tk.W, padx=4)
        log_frame = ttk.Frame(prog)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        self.log = tk.Text(log_frame, height=12, wrap=tk.NONE,
                            font=("Consolas", 9))
        self.log.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb = ttk.Scrollbar(log_frame, command=self.log.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.log.config(yscrollcommand=sb.set)
        self.log.tag_config("ok", foreground="#1a7e2c")
        self.log.tag_config("review", foreground="#cc7000")
        self.log.tag_config("error", foreground="#cc0000")

    def _pick_input(self):
        path = filedialog.askopenfilename(
            title="Select question bank file",
            filetypes=[
                ("Question banks", "*.docx *.xlsx *.csv"),
                ("Word", "*.docx"), ("Excel", "*.xlsx"), ("CSV", "*.csv"),
                ("All files", "*.*"),
            ],
            initialdir=(os.path.dirname(self.input_file.get())
                        if self.input_file.get() else str(Path.home())),
        )
        if path:
            self.input_file.set(path)
            self.settings.set("shuffler_input_file", path)
            self.settings.save()

    def _pick_output(self):
        path = filedialog.askdirectory(
            title="Select output folder for generated sets",
            initialdir=self.output_dir.get() or str(Path.home()),
        )
        if path:
            self.output_dir.set(path)
            self.settings.set("shuffler_output_dir", path)
            self.settings.save()

    def _log(self, text: str, tag: str = "") -> None:
        if tag:
            self.log.insert(tk.END, text + "\n", tag)
        else:
            self.log.insert(tk.END, text + "\n")
        self.log.see(tk.END)

    def _start(self):
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showwarning("Already running",
                                    "Generation already in progress.")
            return
        inp = self.input_file.get().strip()
        if not inp or not os.path.isfile(inp):
            messagebox.showerror("Input",
                                  "Please choose a valid input file.")
            return
        if Path(inp).suffix.lower() not in self.SUPPORTED_INPUTS:
            messagebox.showerror(
                "Input",
                f"Unsupported file type. Use one of: "
                f"{', '.join(self.SUPPORTED_INPUTS)}")
            return
        out = self.output_dir.get().strip()
        if not out:
            messagebox.showerror("Output",
                                  "Please choose an output folder.")
            return

        # Save settings
        self.settings.set("shuffler_n_sets", self.n_sets.get())
        self.settings.set("shuffler_layout", self.layout.get())
        self.settings.set("shuffler_output_format", self.output_format.get())
        self.settings.set("shuffler_shuffle_options", self.shuffle_options.get())
        self.settings.save()

        self.log.delete("1.0", tk.END)
        self._log(f"Generating {self.n_sets.get()} sets from "
                  f"{os.path.basename(inp)}…")
        if self.output_format.get() == "pdf":
            if pdf_engine_available():
                self._log(f"  PDF engine: {pdf_engine_name()}")
            else:
                self._log("  ⚠ PDF engine not found — will keep .docx output", "review")

        self.cancel_event.clear()
        self.progress_var.set(0)
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)

        job = ShuffleJob(
            input_file=Path(inp),
            output_dir=Path(out),
            n_sets=self.n_sets.get(),
            layout=self.layout.get(),
            output_format=self.output_format.get(),
            shuffle_options=self.shuffle_options.get(),
            title_prefix=self.title_prefix.get(),
            progress_queue=self.progress_queue,
            cancel_event=self.cancel_event,
        )
        self.worker_thread = threading.Thread(target=job.run, daemon=True)
        self.worker_thread.start()

    def _stop(self):
        if messagebox.askyesno("Stop", "Stop generation?"):
            self.cancel_event.set()
            self._log("Cancelling…", "error")

    def _clear(self):
        """Reset the UI for a fresh run."""
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showwarning("Busy", "Generation is running. Stop it first.")
            return
        self.log.delete("1.0", tk.END)
        self.progress_var.set(0)
        self.status_label.config(text="Ready.")
        self._log("Cleared. Ready to generate sets.")

    def _poll_queue(self):
        try:
            while True:
                msg = self.progress_queue.get_nowait()
                self._handle_msg(msg)
        except queue.Empty:
            pass
        self.after(100, self._poll_queue)

    def _handle_msg(self, msg):
        kind = msg.get("type")
        if kind == "start":
            self.status_label.config(text=f"Starting… {msg['total']} sets")
        elif kind == "log":
            self._log("  " + msg["text"])
        elif kind == "progress":
            pct = msg["done"] * 100 / msg["total"]
            self.progress_var.set(pct)
            self.status_label.config(
                text=f"{msg['done']}/{msg['total']} ({pct:.0f}%)")
            self._log(f"  ✓ {msg['current']} written", "ok")
        elif kind == "done":
            self.progress_var.set(100)
            self.status_label.config(
                text=f"Done in {msg['elapsed_sec']:.1f}s")
            self._log(f"\n✓ All sets generated in "
                      f"{msg['elapsed_sec']:.1f}s", "ok")
            self.start_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)
            messagebox.showinfo(
                "Done",
                f"Sets generated.\n\nOutput folder: {self.output_dir.get()}",
            )
        elif kind == "cancelled":
            self.status_label.config(text="Cancelled.")
            self.start_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)
        elif kind == "error":
            self.status_label.config(text="Error.")
            self._log(f"ERROR: {msg['message']}", "error")
            if msg.get("traceback"):
                self._log(msg["traceback"], "error")
            self.start_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)


# ============================================================================
# Main window
# ============================================================================


class MCQStudio:
    def __init__(self) -> None:
        self.settings = Settings()
        # Make sure default folders exist on first run
        ensure_dir(self.settings.get("omr_output_dir"))
        ensure_dir(self.settings.get("shuffler_output_dir"))

        self.root = tk.Tk()
        self.root.title(f"{APP_NAME} {APP_VERSION}")
        self.root.geometry("840x640")
        self.root.minsize(720, 540)

        # Use nicer ttk theme where available
        try:
            style = ttk.Style()
            for theme in ("vista", "winnative", "clam"):
                if theme in style.theme_names():
                    style.theme_use(theme)
                    break
        except Exception:
            pass

        self._build_menu()
        self._build_notebook()
        self._build_statusbar()

    def _build_menu(self):
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        filem = tk.Menu(menubar, tearoff=0)
        filem.add_command(label="Open OMR Output Folder",
                          command=lambda: open_folder(
                              self.settings.get("omr_output_dir")))
        filem.add_command(label="Open Shuffler Output Folder",
                          command=lambda: open_folder(
                              self.settings.get("shuffler_output_dir")))
        filem.add_separator()
        filem.add_command(label="Exit", command=self.root.destroy)
        menubar.add_cascade(label="File", menu=filem)

        helpm = tk.Menu(menubar, tearoff=0)
        helpm.add_command(label="About", command=self._show_about)
        menubar.add_cascade(label="Help", menu=helpm)

    def _build_notebook(self):
        nb = ttk.Notebook(self.root)
        nb.pack(fill=tk.BOTH, expand=True, padx=8, pady=(8, 4))

        self.omr_tab = OmrTab(nb, self.settings)
        self.shuf_tab = ShufflerTab(nb, self.settings)
        nb.add(self.omr_tab, text="  📊  OMR Scanner  ")
        nb.add(self.shuf_tab, text="  🔀  MCQ Shuffler  ")

    def _build_statusbar(self):
        bar = ttk.Frame(self.root, relief=tk.SUNKEN, padding=4)
        bar.pack(side=tk.BOTTOM, fill=tk.X)
        pdf_info = f"PDF: {pdf_engine_name()}" if pdf_engine_available() else "PDF: install docx2pdf or LibreOffice"
        ttk.Label(
            bar,
            text=f"{APP_NAME} {APP_VERSION}  •  {pdf_info}",
            font=("Segoe UI", 8),
        ).pack(side=tk.LEFT)

    def _show_about(self):
        messagebox.showinfo(
            f"About {APP_NAME}",
            f"{APP_NAME}  v{APP_VERSION}\n\n"
            f"Developer:  {APP_DEVELOPER}\n"
            f"Company:    {APP_COMPANY}\n\n"
            "A unified desktop tool for OMR scanning\n"
            "and MCQ paper generation.\n\n"
            f"PDF engine: {pdf_engine_name()}\n"
            f"Python: {sys.version.split()[0]}\n"
            f"Platform: {sys.platform}",
        )

    def run(self):
        self.root.mainloop()


def main() -> int:
    try:
        app = MCQStudio()
        app.run()
        return 0
    except Exception as e:
        try:
            messagebox.showerror(
                f"{APP_NAME} — Crash",
                f"{type(e).__name__}: {e}\n\n{traceback.format_exc()}",
            )
        except Exception:
            traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
